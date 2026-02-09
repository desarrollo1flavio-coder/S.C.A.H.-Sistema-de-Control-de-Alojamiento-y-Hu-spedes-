"""Parser de archivos Excel para S.C.A.H.

Lee, mapea columnas automáticamente, valida filas y retorna datos
listos para importar a la base de datos.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from config.settings import COLUMNAS_IGNORAR, COLUMNAS_MAPEO, MAX_IMPORT_ROWS, PREVIEW_ROWS
from utils.logger import get_logger
from utils.validators import sanitizar_texto

logger = get_logger("utils.excel_parser")


class ExcelParser:
    """Parser de archivos Excel con mapeo automático de columnas."""

    def __init__(self, file_path: Path, selected_sheets: Optional[list[str]] = None) -> None:
        """Inicializa el parser.

        Args:
            file_path: Ruta al archivo Excel.
            selected_sheets: Lista de nombres de hojas a procesar.
                Si es None, se procesan todas las hojas numéricas.
        """
        self._file_path = Path(file_path)
        self._selected_sheets = selected_sheets
        self._column_map: dict[str, str] = {}
        self._sheet_count: int = 0
        self._sheet_names: list[str] = []
        self._raw_frames: list[pd.DataFrame] = []

    @staticmethod
    def list_sheets(file_path: Path, only_numeric: bool = False) -> list[str]:
        """Lista las hojas disponibles en un archivo Excel.

        Args:
            file_path: Ruta al archivo Excel.
            only_numeric: Si True, solo retorna hojas con nombres numéricos.

        Returns:
            Lista de nombres de hojas.
        """
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()
        engine = "openpyxl" if suffix == ".xlsx" else "xlrd"

        if engine == "xlrd":
            import xlrd
            wb = xlrd.open_workbook(str(file_path), on_demand=True)
            all_names = wb.sheet_names()
            wb.release_resources()
        else:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), read_only=True)
            all_names = wb.sheetnames
            wb.close()

        if only_numeric:
            # Filtrar solo hojas con nombre numérico
            result = [n for n in all_names if str(n).strip().isdigit()]
            # Ordenar numéricamente
            result.sort(key=lambda x: int(str(x).strip()))
            return result

        # Retornar todas las hojas en orden original
        return list(all_names)

    def parse(self) -> dict:
        """Lee y procesa el archivo Excel.

        Returns:
            Diccionario con:
                - total_rows: número total de filas leídas
                - valid_rows: lista de dicts con datos válidos
                - errors: lista de dicts {row, error}
                - duplicates: lista de dicts con filas duplicadas
                - column_mapping: mapeo de columnas detectado
        """
        logger.info("Parseando archivo: %s", self._file_path)

        df = self._read_file()
        if df.empty:
            return {
                "total_rows": 0, "valid_rows": [], "errors": [],
                "duplicates": [], "column_mapping": {},
            }

        # Limitar filas
        total_rows = len(df)
        if total_rows > MAX_IMPORT_ROWS:
            logger.warning(
                "Archivo truncado: %d filas (máx %d)", total_rows, MAX_IMPORT_ROWS,
            )
            df = df.head(MAX_IMPORT_ROWS)

        # Columnas ya mapeadas por hoja en _read_file; solo log
        logger.info("Mapeo de columnas consolidado: %s", self._column_map)

        # Procesar filas
        valid_rows: list[dict] = []
        errors: list[dict] = []
        seen_docs: set[str] = set()
        duplicates: list[dict] = []
        skipped: int = 0

        for idx, row in df.iterrows():
            hoja = row.get("_hoja_origen", "")
            fila = row.get("_fila_original", int(idx) + 2)
            row_label = f"Hoja '{hoja}' fila {fila}" if hoja else f"Fila {fila}"
            try:
                data = self._process_row(row)
                if not data:
                    skipped += 1
                    continue

                # Guardar hoja de origen en el registro
                if hoja:
                    data["_hoja_origen"] = hoja

                # Verificar duplicados dentro del archivo
                doc_key = data.get("dni") or data.get("pasaporte") or ""
                if doc_key and doc_key in seen_docs:
                    duplicates.append({"row": row_label, **data})
                    continue
                if doc_key:
                    seen_docs.add(doc_key)

                valid_rows.append(data)

            except Exception as e:
                errors.append({"row": row_label, "error": str(e)})

        # Construir raw_preview (datos crudos del Excel para la vista previa)
        # Muestreo round-robin: 1 fila por hoja para mostrar variedad
        raw_preview: list[dict] = []
        if self._raw_frames:
            try:
                # Tomar 1 fila válida de cada hoja en round-robin
                sampled: list[pd.DataFrame] = []
                for rf in self._raw_frames:
                    # Filtrar filas vacías (excluir columna "Hoja" del chequeo)
                    data_cols = [c for c in rf.columns if c != "Hoja"]
                    valid_mask = rf[data_cols].astype(str).apply(
                        lambda x: x.str.strip().str.len().sum(), axis=1,
                    ) > 0
                    valid_rows_rf = rf[valid_mask]
                    if not valid_rows_rf.empty:
                        sampled.append(valid_rows_rf.head(1))

                if sampled:
                    sample_df = pd.concat(sampled, ignore_index=True, sort=False)
                    sample_df = sample_df.fillna("")
                    raw_preview = sample_df.head(PREVIEW_ROWS).to_dict(orient="records")
            except Exception:
                raw_preview = []

        result = {
            "total_rows": total_rows,
            "valid_rows": valid_rows,
            "errors": errors,
            "duplicates": duplicates,
            "skipped": skipped,
            "column_mapping": self._column_map,
            "sheet_count": self._sheet_count,
            "sheet_names": self._sheet_names,
            "raw_preview": raw_preview,
        }

        logger.info(
            "Resultado: %d válidos, %d errores, %d duplicados, %d omitidas de %d totales",
            len(valid_rows), len(errors), len(duplicates), skipped, total_rows,
        )

        return result

    def _read_file(self) -> pd.DataFrame:
        """Lee TODAS las hojas del archivo Excel, mapea columnas por hoja y las concatena.

        Cada hoja recibe su propio mapeo de columnas para que hojas con
        nombres de columna diferentes se procesen correctamente.

        Returns:
            DataFrame con datos combinados de todas las hojas, columnas
            ya renombradas a los campos del sistema.
        """
        try:
            suffix = self._file_path.suffix.lower()
            engine = "openpyxl" if suffix == ".xlsx" else "xlrd"

            # Leer TODAS las hojas (sheet_name=None devuelve dict)
            all_sheets = pd.read_excel(
                self._file_path,
                engine=engine,
                dtype=str,
                keep_default_na=False,
                sheet_name=None,
            )

            self._sheet_count = len(all_sheets)
            self._sheet_names = list(all_sheets.keys())
            logger.info(
                "Hojas encontradas (%d): %s",
                self._sheet_count, self._sheet_names,
            )

            # Determinar qué hojas procesar
            if self._selected_sheets is not None:
                # El usuario seleccionó hojas específicas: usar esas
                sel_set = {str(s).strip() for s in self._selected_sheets}
                hojas_a_procesar = {}
                for name, df_sheet in all_sheets.items():
                    if str(name).strip() in sel_set:
                        hojas_a_procesar[name] = df_sheet
                logger.info(
                    "Hojas seleccionadas por usuario: %s de %d disponibles",
                    list(sel_set), len(all_sheets),
                )
            else:
                # Sin selección: procesar todas las hojas
                hojas_a_procesar = dict(all_sheets)
                logger.info("Procesando todas las hojas (%d)", len(hojas_a_procesar))

            all_mappings: dict[str, str] = {}
            frames: list[pd.DataFrame] = []
            self._raw_frames = []

            for sheet_name, sheet_df in hojas_a_procesar.items():
                if sheet_df.empty:
                    logger.info("Hoja '%s' vacía, omitida", sheet_name)
                    continue

                # Limpiar nombres de columnas
                sheet_df.columns = [
                    str(col).strip().lower().replace(" ", "_")
                    for col in sheet_df.columns
                ]

                # Eliminar filas completamente vacías
                sheet_df = sheet_df.dropna(how="all")
                sheet_df = sheet_df[
                    sheet_df.astype(str).apply(
                        lambda x: x.str.strip().str.len().sum(), axis=1,
                    ) > 0
                ]

                if sheet_df.empty:
                    continue

                # Mapear columnas POR HOJA para manejar nombres diferentes
                sheet_cols = [c for c in sheet_df.columns.tolist() if not c.startswith("_")]
                sheet_map = self._auto_map_columns(sheet_cols)
                all_mappings.update(sheet_map)

                # Guardar copia RAW con headers legibles (Title Case) para preview
                import unicodedata

                def _normalize_ignore(text: str) -> str:
                    t = text.strip().lower().replace(" ", "_")
                    nfkd = unicodedata.normalize("NFKD", t)
                    return nfkd.encode("ascii", "ignore").decode("ascii").replace(".", "").replace("°", "").replace("º", "")

                ignorar_norm = [_normalize_ignore(ig) for ig in COLUMNAS_IGNORAR]
                # Columnas con nombre original legible, excluyendo las ignoradas y las internas
                raw_cols = []
                for c in sheet_df.columns:
                    if c.startswith("_"):
                        continue
                    if _normalize_ignore(c) in ignorar_norm:
                        continue
                    raw_cols.append(c)
                raw_copy = sheet_df[raw_cols].copy()
                # Renombrar a Title Case legible: "apellido_y_nombre" → "Apellido Y Nombre"
                raw_copy.columns = [
                    str(c).replace("_", " ").strip().title() for c in raw_copy.columns
                ]
                # Agregar columna "Hoja" con el número de hoja de origen
                raw_copy.insert(0, "Hoja", str(sheet_name))
                self._raw_frames.append(raw_copy)

                sheet_df = sheet_df.rename(columns=sheet_map)

                sheet_df = sheet_df.copy()
                sheet_df["_hoja_origen"] = str(sheet_name)
                # Guardar número de fila original dentro de la hoja
                sheet_df["_fila_original"] = range(2, len(sheet_df) + 2)
                frames.append(sheet_df)
                logger.info(
                    "Hoja '%s': %d filas, %d columnas mapeadas",
                    sheet_name, len(sheet_df), len(sheet_map),
                )

            self._column_map = all_mappings
            self._sheet_count = len(frames)  # Solo hojas numéricas con datos
            self._sheet_names = [str(f["_hoja_origen"].iloc[0]) for f in frames if len(f) > 0]

            if not frames:
                logger.warning("Ninguna hoja contiene datos")
                return pd.DataFrame()

            # Concatenar todas las hojas
            df = pd.concat(frames, ignore_index=True, sort=False)
            # Rellenar NaN de columnas faltantes con cadena vacía
            df = df.fillna("")

            logger.info(
                "Total combinado: %d filas de %d hojas",
                len(df), len(frames),
            )
            return df

        except Exception as e:
            logger.error("Error al leer archivo Excel: %s", e)
            raise ValueError(f"No se pudo leer el archivo: {e}") from e

    def _auto_map_columns(self, file_columns: list[str]) -> dict[str, str]:
        """Mapea columnas del archivo a campos del sistema.

        Args:
            file_columns: Nombres de columnas del archivo.

        Returns:
            Mapeo {columna_archivo: campo_sistema}.
        """
        import unicodedata

        def normalize(text: str) -> str:
            """Normaliza texto removiendo acentos y caracteres especiales."""
            text = text.strip().lower().replace(" ", "_")
            nfkd = unicodedata.normalize("NFKD", text)
            ascii_text = nfkd.encode("ascii", "ignore").decode("ascii")
            ascii_text = ascii_text.replace(".", "").replace("°", "").replace("º", "")
            return ascii_text

        mapping: dict[str, str] = {}
        used_fields: set[str] = set()

        # Normalizar lista de columnas a ignorar
        ignorar_norm = [normalize(ig) for ig in COLUMNAS_IGNORAR]

        # Filtrar columnas que son contadores/numeradores
        cols_validas = []
        for col in file_columns:
            col_norm = normalize(col)
            if col_norm in ignorar_norm:
                logger.debug("Columna ignorada (contador): %s", col)
            else:
                cols_validas.append(col)

        for system_field, aliases in COLUMNAS_MAPEO.items():
            normalized_aliases = [normalize(a) for a in aliases]
            for col in cols_validas:
                if col in mapping:
                    continue
                col_clean = col.strip().lower().replace(" ", "_")
                col_normalized = normalize(col)

                # Coincidencia exacta o normalizada
                if col_clean in aliases or col_normalized in normalized_aliases:
                    mapping[col] = system_field
                    used_fields.add(system_field)
                    break

                # Coincidencia parcial (solo con aliases de 5+ caracteres)
                for alias in aliases:
                    alias_norm = normalize(alias)
                    if len(alias_norm) >= 5 and (
                        alias_norm in col_normalized or col_normalized in alias_norm
                    ):
                        if system_field not in used_fields:
                            mapping[col] = system_field
                            used_fields.add(system_field)
                            break
                if system_field in used_fields:
                    break

        logger.info("Columnas del archivo: %s", file_columns)
        logger.info("Mapeo resultante: %s", mapping)
        logger.info("Columnas no mapeadas: %s", [c for c in file_columns if c not in mapping])

        return mapping

    def _process_row(self, row: pd.Series) -> Optional[dict]:
        """Procesa una fila del DataFrame.

        Args:
            row: Serie de Pandas con datos de la fila.

        Returns:
            Diccionario con datos procesados o None si la fila es inválida.
        """
        data: dict[str, Any] = {}

        # Campo combinado "APELLIDO Y NOMBRE" → dividir en apellido + nombre
        apellido_nombre_val = self._get_value(row, "apellido_nombre")
        if apellido_nombre_val:
            texto = apellido_nombre_val.strip()
            if "," in texto:
                # Formato: "GONZALEZ, JUAN CARLOS"
                partes = texto.split(",", 1)
                data["apellido"] = sanitizar_texto(partes[0].strip())
                data["nombre"] = sanitizar_texto(partes[1].strip()) if len(partes) > 1 and partes[1].strip() else ""
            else:
                # Formato: "GONZALEZ JUAN CARLOS" — primera palabra = apellido
                partes = texto.split(None, 1)
                data["apellido"] = sanitizar_texto(partes[0]) if partes else ""
                data["nombre"] = sanitizar_texto(partes[1]) if len(partes) > 1 else ""

        # Campos de texto
        for field in ("apellido", "nombre", "nacionalidad", "procedencia",
                       "profesion", "establecimiento", "habitacion", "destino",
                       "telefono", "vehiculo"):
            # No sobrescribir apellido/nombre si ya se extrajeron del campo combinado
            if field in ("apellido", "nombre") and field in data and data[field]:
                continue
            val = self._get_value(row, field)
            if val:
                data[field] = sanitizar_texto(val) if field in (
                    "apellido", "nombre", "nacionalidad", "procedencia",
                ) else val.strip()

        # Si no tiene apellido ni nombre, no es una fila de datos reales
        if "apellido" not in data and "nombre" not in data:
            return None

        # Documentos — detectar tipo automáticamente
        dni = self._get_value(row, "dni")
        pasaporte = self._get_value(row, "pasaporte")

        # Si hay columna combinada DNI/Pasaporte, detectar tipo
        if dni and not pasaporte:
            cleaned = dni.strip().replace(".", "").replace("-", "").replace(" ", "")
            if cleaned.endswith(".0"):  # Fix float de Excel
                cleaned = cleaned[:-2]
            if cleaned.isdigit() and 7 <= len(cleaned) <= 8:
                data["dni"] = cleaned
            elif cleaned.isalnum() and 5 <= len(cleaned) <= 15:
                data["pasaporte"] = cleaned.upper()
        else:
            if dni:
                cleaned = dni.strip().replace(".", "").replace("-", "").replace(" ", "")
                if cleaned.endswith(".0"):
                    cleaned = cleaned[:-2]
                if cleaned.isdigit() and 7 <= len(cleaned) <= 8:
                    data["dni"] = cleaned
            if pasaporte:
                cleaned_p = pasaporte.strip().upper().replace(" ", "")
                if cleaned_p.isalnum() and 5 <= len(cleaned_p) <= 15:
                    data["pasaporte"] = cleaned_p

        # Fallback: buscar número de 7-8 dígitos en cualquier columna
        if "dni" not in data and "pasaporte" not in data:
            for col_name, col_val in row.items():
                if str(col_name).startswith("_"):
                    continue
                val = str(col_val).strip().replace(".", "").replace("-", "").replace(" ", "")
                if val.endswith(".0"):
                    val = val[:-2]
                if val.isdigit() and 7 <= len(val) <= 8:
                    data["dni"] = val
                    break

        # Validar documento — solo reportar error si tiene AMBOS nombres
        if "dni" not in data and "pasaporte" not in data:
            if "apellido" in data and "nombre" in data:
                raise ValueError("Sin documento válido (DNI o Pasaporte)")
            return None  # Fila parcial sin nombres completos ni doc: omitir

        # Validar campos obligatorios
        for required in ("apellido", "nombre"):
            if required not in data or not data[required]:
                raise ValueError(f"Campo obligatorio '{required}' vacío")

        # Defaults para campos obligatorios faltantes
        if "nacionalidad" not in data:
            data["nacionalidad"] = "Argentina"
        if "procedencia" not in data:
            data["procedencia"] = "Sin especificar"
        if "habitacion" not in data:
            data["habitacion"] = "S/N"

        # Edad
        edad = self._get_value(row, "edad")
        if edad:
            try:
                edad_int = int(float(edad))
                if 0 < edad_int < 150:
                    data["edad"] = edad_int
            except (ValueError, TypeError):
                pass

        # Fecha de nacimiento
        fecha_nac = self._get_value(row, "fecha_nacimiento")
        if fecha_nac:
            parsed_nac = self._parse_date(fecha_nac)
            if parsed_nac:
                data["fecha_nacimiento"] = parsed_nac

        # Fechas
        data["fecha_entrada"] = self._parse_date(
            self._get_value(row, "fecha_entrada"),
        ) or date.today()

        fecha_salida = self._parse_date(self._get_value(row, "fecha_salida"))
        if fecha_salida:
            data["fecha_salida"] = fecha_salida

        # Vehículo
        vehiculo = data.pop("vehiculo", None)
        if vehiculo:
            data["vehiculo_tiene"] = True
            data["vehiculo_datos"] = vehiculo
        else:
            data["vehiculo_tiene"] = False

        return data

    @staticmethod
    def _get_value(row: pd.Series, field: str) -> Optional[str]:
        """Obtiene un valor de la fila.

        Args:
            row: Serie de datos.
            field: Nombre del campo.

        Returns:
            Valor como string o None.
        """
        if field in row.index:
            val = row[field]
            if pd.notna(val) and str(val).strip():
                return str(val).strip()
        return None

    @staticmethod
    def _parse_date(value: Optional[str]) -> Optional[date]:
        """Parsea una fecha desde múltiples formatos.

        Args:
            value: String con la fecha.

        Returns:
            Objeto date o None.
        """
        if not value:
            return None

        formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%d.%m.%Y",
            "%m/%d/%Y",
            "%Y/%m/%d",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue

        # Intentar con pandas
        try:
            return pd.to_datetime(value).date()
        except Exception:
            return None
