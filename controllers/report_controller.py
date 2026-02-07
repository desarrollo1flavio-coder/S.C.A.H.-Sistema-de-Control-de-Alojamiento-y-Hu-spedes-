"""Controlador de reportes para S.C.A.H.

Genera exportaciones en formato Excel y PDF con formato institucional.
"""

import io
from datetime import datetime
from pathlib import Path
from typing import Optional

from controllers.auth_controller import SessionInfo
from models.auditoria import AuditoriaDAO
from models.huesped import HuespedDAO
from utils.exceptions import PermissionDeniedError
from utils.logger import get_logger

logger = get_logger("report_controller")


class ReportController:
    """Controlador para generación de reportes y exportaciones."""

    def __init__(self, session: SessionInfo) -> None:
        """Inicializa el controlador.

        Args:
            session: Sesión del usuario autenticado.
        """
        self._session = session

    def export_excel(
        self,
        data: list[dict],
        output_path: str | Path,
        *,
        title: str = "Reporte de Huéspedes",
    ) -> Path:
        """Exporta datos a un archivo Excel formateado.

        Args:
            data: Lista de registros a exportar.
            output_path: Ruta de salida del archivo.
            title: Título del reporte.

        Returns:
            Path del archivo generado.

        Raises:
            PermissionDeniedError: Sin permisos de lectura.
        """
        if not self._session.tiene_permiso("lectura"):
            raise PermissionDeniedError("Sin permisos para generar reportes")

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Huéspedes"

        # Estilos
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        title_font = Font(name="Calibri", bold=True, size=14)
        subtitle_font = Font(name="Calibri", size=10, italic=True, color="666666")

        # Título institucional
        ws.merge_cells("A1:N1")
        ws["A1"] = "Policía de Tucumán — Departamento de Inteligencia Criminal — Sección Hoteles"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12.5, color="1F4E79")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:N2")
        ws["A2"] = title
        ws["A2"].font = title_font
        ws["A2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A3:N3")
        ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} — Usuario: {self._session.nombre_completo}"
        ws["A3"].font = subtitle_font
        ws["A3"].alignment = Alignment(horizontal="center")

        # Columnas
        columns = [
            ("ID", 8), ("Apellido", 18), ("Nombre", 18),
            ("DNI", 12), ("Pasaporte", 12), ("Nacionalidad", 15),
            ("Procedencia", 18), ("Edad", 8), ("Profesión", 15),
            ("Habitación", 12), ("Entrada", 12), ("Salida", 12),
            ("Teléfono", 14), ("Destino", 18),
        ]

        start_row = 5
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Datos
        field_keys = [
            "id", "apellido", "nombre", "dni", "pasaporte",
            "nacionalidad", "procedencia", "edad", "profesion",
            "habitacion", "fecha_entrada", "fecha_salida",
            "telefono", "destino",
        ]

        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
        data_align = Alignment(vertical="center")

        for row_idx, record in enumerate(data, start_row + 1):
            for col_idx, key in enumerate(field_keys, 1):
                val = record.get(key, "")
                if val is None:
                    val = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Pie
        last_row = start_row + len(data) + 2
        ws.merge_cells(f"A{last_row}:N{last_row}")
        ws[f"A{last_row}"] = f"Total de registros: {len(data)}"
        ws[f"A{last_row}"].font = Font(name="Calibri", bold=True, size=10)

        output = Path(output_path)
        wb.save(output)

        # Auditoría
        AuditoriaDAO.registrar(
            usuario=self._session.username,
            accion="SELECT",
            tabla_afectada="huespedes",
            detalles=f"Exportación Excel: {len(data)} registros — {output.name}",
        )

        logger.info("Excel exportado: %s (%d registros)", output, len(data))
        return output

    def export_pdf(
        self,
        data: list[dict],
        output_path: str | Path,
        *,
        title: str = "Reporte de Huéspedes",
    ) -> Path:
        """Exporta datos a un archivo PDF con formato institucional.

        Args:
            data: Lista de registros.
            output_path: Ruta de salida.
            title: Título del reporte.

        Returns:
            Path del archivo generado.
        """
        if not self._session.tiene_permiso("lectura"):
            raise PermissionDeniedError("Sin permisos para generar reportes")

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        output = Path(output_path)
        doc = SimpleDocTemplate(
            str(output),
            pagesize=landscape(A4),
            leftMargin=1.5 * cm,
            rightMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Header
        header_style = ParagraphStyle(
            "Header", parent=styles["Title"],
            fontSize=13, textColor=colors.HexColor("#1F4E79"),
            spaceAfter=3 * mm,
        )
        elements.append(
            Paragraph("Policía de Tucumán — Sección Hoteles", header_style),
        )
        elements.append(
            Paragraph(title, ParagraphStyle(
                "Subtitle", parent=styles["Title"], fontSize=14, spaceAfter=2 * mm,
            )),
        )
        elements.append(
            Paragraph(
                f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')} — "
                f"Usuario: {self._session.nombre_completo}",
                ParagraphStyle(
                    "Info", parent=styles["Normal"],
                    fontSize=8, textColor=colors.grey, spaceAfter=5 * mm,
                ),
            ),
        )

        # Tabla
        headers = [
            "Apellido", "Nombre", "DNI", "Pasaporte",
            "Nacionalidad", "Hab.", "Entrada", "Salida", "Teléfono",
        ]
        field_keys = [
            "apellido", "nombre", "dni", "pasaporte",
            "nacionalidad", "habitacion", "fecha_entrada", "fecha_salida",
            "telefono",
        ]

        table_data = [headers]
        for record in data:
            row = []
            for key in field_keys:
                val = record.get(key, "")
                row.append(str(val) if val else "")
            table_data.append(row)

        col_widths = [80, 80, 60, 65, 70, 35, 65, 65, 70]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7FB")]),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 10 * mm))
        elements.append(
            Paragraph(
                f"Total de registros: {len(data)}",
                ParagraphStyle(
                    "Total", parent=styles["Normal"],
                    fontSize=9, fontName="Helvetica-Bold",
                ),
            ),
        )

        doc.build(elements)

        # Auditoría
        AuditoriaDAO.registrar(
            usuario=self._session.username,
            accion="SELECT",
            tabla_afectada="huespedes",
            detalles=f"Exportación PDF: {len(data)} registros — {output.name}",
        )

        logger.info("PDF exportado: %s (%d registros)", output, len(data))
        return output
